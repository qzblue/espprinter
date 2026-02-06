from __future__ import annotations

from datetime import datetime
from io import BytesIO
from typing import Any, Dict, List, Optional, Tuple
from urllib.parse import urlparse
import urllib.parse
from collections import defaultdict
import subprocess

from flask import Flask, render_template, request, send_file, Response, stream_with_context
from flask_caching import Cache

try:
    from openpyxl import Workbook
except ImportError as exc:  # pragma: no cover - runtime guard
    raise RuntimeError("請先安裝 openpyxl 套件：pip install openpyxl") from exc

from sharp_mfp_export import (
    PRINTERS,
    DEFAULT_USAGE_CATEGORIES,
    USAGE_CATEGORY_CONFIG,
    aggregate_joblog_reports,
    aggregate_usage_by_categories,
    format_dt,
    load_joblog_report,
    parse_month_range,
    parse_time_value,
    parse_week_range,
    fetch_latest_user_counts,
    fetch_aggregated_users_paginated,
    fetch_job_logs_by_users,
    host_tag,
    normalize_name,
    run_download_process,
    fetch_total_user_printer_pairs,
    log_update_event,
    fetch_total_jobs_count,
)

import logging
import ldap_service

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)

app = Flask(__name__)

# Configure caching with simple in-memory cache
app.config['CACHE_TYPE'] = 'SimpleCache'
app.config['CACHE_DEFAULT_TIMEOUT'] = 300  # 5 minutes
cache = Cache(app)

# Register custom Jinja2 filter for printer label conversion
@app.template_filter('printer_label')
def printer_label_filter(url: str) -> str:
    """Convert printer URL to friendly label for use in templates"""
    return _printer_label(url)

# Register custom Jinja2 filter for user display name from LDAP
@app.template_filter('user_display_name')
def user_display_name_filter(username: str, show_username: bool = True) -> str:
    """Convert username to display name from Active Directory"""
    return ldap_service.format_user_display(username, show_username)


try:
    from sharp_mfp_export import PRINTER_ALIASES
except ImportError:
    PRINTER_ALIASES = {}

def _printer_label(url: str) -> str:
    if url in PRINTER_ALIASES:
        return PRINTER_ALIASES[url]
    parsed = urlparse(url)
    return parsed.netloc or url


PRINTER_CHOICES = [{"value": printer, "label": _printer_label(printer)} for printer in PRINTERS]
CATEGORY_CHOICES = [
    {"value": key, "label": config["label"]} for key, config in USAGE_CATEGORY_CONFIG.items()
]


def _selected_printers(choice: Optional[str]) -> List[str]:
    if not choice or choice == "all":
        return PRINTERS
    return [choice]


def _to_int(value: Optional[str], default: int) -> int:
    try:
        return int(value)
    except (TypeError, ValueError):
        return default


def _parse_query_datetime(raw: str, label: str, errors: List[str]) -> Optional[datetime]:
    if not raw:
        return None
    parsed = parse_time_value(raw)
    if not parsed:
        errors.append(f"{label} 格式無法解析：{raw}")
    return parsed




def _parse_common_params(query_args: Dict[str, str]) -> Dict[str, Any]:
    """Extract and normalize common query parameters."""
    time_mode = query_args.get("time_mode", "all")
    
    # Defaults and simple cleanups
    params = {
        "printer": query_args.get("printer", "all"),
        "user": query_args.get("user", "").strip(),
        "mode": query_args.get("mode", "").strip(),
        "computer": query_args.get("computer", "").strip(),
        "filename": query_args.get("filename", "").strip(),
        "time_mode": time_mode,
        "limit": _to_int(query_args.get("limit", "5"), 5),
        "page": _to_int(query_args.get("page", "1"), 1),
        "per_page": _to_int(query_args.get("per_page", "0"), 0), # Default depends on context, set 0 here
        "month": "",
        "week": "",
        "start": "",
        "end": "",
    }
    
    # Handle Time Mode Logic
    if time_mode != "all":
        params["month"] = query_args.get("month", "").strip()
        params["week"] = query_args.get("week", "").strip()
        params["start"] = query_args.get("start", "").strip()
        params["end"] = query_args.get("end", "").strip()
        
    return params

from threading import Lock
import threading

download_lock = Lock()

@app.route("/update_data")
def update_data():
    def generate():
        token = request.args.get("token", "")
        expected_token = "2851@9364"

        if token != expected_token:
            yield 'data: {"status": "error", "message": "密碼錯誤，您沒有權限執行更新。"}\n\n'
            return

        # Try to acquire lock
        if not download_lock.acquire(blocking=False):
            yield 'data: {"status": "error", "message": "已有更新正在進行中，請稍後再試。"}\n\n'
            return

        # Init Log
        log_id = 0
        try:
            log_id = log_update_event("web_manual", "running", "開始更新程序 (Web)", 0)
            yield 'data: {"status": "start", "message": "開始更新程序..."}\n\n'

            # Use internal generator instead of subprocess
            for line in run_download_process(None): # None = all printers
                line = line.strip()
                if not line:
                    continue
                if line.startswith("== http"):
                    yield f'data: {{"status": "progress", "message": "正在處理: {line[3:]}"}}\n\n'
                else:
                    yield f'data: {{"status": "log", "message": "{line}"}}\n\n'

            # If we completed the loop, it means success (exceptions yielded as FAIL messages)
            
            yield 'data: {"status": "log", "message": "正在預熱緩存..."}\n\n'
            try:
                # Warm up cache for main endpoints
                with app.test_client() as client:
                    client.get("/counts")
                    client.get("/jobs")
                    client.get("/leaders")
                
                log_update_event("web_manual", "success", "更新完成 (Web)", log_id)
                yield 'data: {"status": "done", "message": "更新完成！"}\n\n'
            except Exception as w_err:
                err_msg = f"緩存預熱失敗: {str(w_err)}"
                log_update_event("web_manual", "warning", err_msg, log_id)
                yield f'data: {{"status": "error", "message": "{err_msg}"}}\n\n'
            
        except Exception as e:
            err_msg = f"系統錯誤: {str(e)}"
            if log_id > 0:
                log_update_event("web_manual", "error", err_msg, log_id)
            yield f'data: {{"status": "error", "message": "{err_msg}"}}\n\n'

        
        finally:
            download_lock.release()

    return Response(stream_with_context(generate()), mimetype="text/event-stream")


@app.route("/.well-known/appspecific/com.chrome.devtools.json")
def chrome_devtools():
    return {}


def _resolve_time_range_from_query(
    mode: str,
    month: str,
    week: str,
    start_raw: str,
    end_raw: str,
    errors: List[str],
) -> Tuple[Optional[datetime], Optional[datetime]]:
    raw_mode = (mode or "all").lower()
    month = month.strip()
    week = week.strip()
    start_raw = start_raw.strip()
    end_raw = end_raw.strip()

    if raw_mode == "all":
        if month:
            mode = "month"
        elif week:
            mode = "week"
        elif start_raw or end_raw:
            mode = "custom"
        else:
            mode = raw_mode
    else:
        mode = raw_mode
    if mode == "month":
        if not month:
            errors.append("請輸入欲查詢的月份 (YYYY-MM)")
            return None, None
        try:
            return parse_month_range(month)
        except ValueError as exc:
            errors.append(str(exc))
            return None, None
    if mode == "week":
        if not week:
            errors.append("請輸入欲查詢的週期 (例如 2026-W05)")
            return None, None
        try:
            return parse_week_range(week)
        except ValueError as exc:
            errors.append(str(exc))
            return None, None
    if mode == "custom":
        start_dt = _parse_query_datetime(start_raw, "開始時間", errors)
        end_dt = _parse_query_datetime(end_raw, "結束時間", errors)
        if start_dt and end_dt and end_dt < start_dt:
            # Smart fix: auto-swap if start > end
            start_dt, end_dt = end_dt, start_dt
            # errors.append("結束時間需晚於開始時間")
            # return None, None
        return start_dt, end_dt

    return None, None


def _collect_job_reports(
    printers: List[str],
    user_kw: Optional[str],
    mode_kw: Optional[str],
    computer_kw: Optional[str],
    start_dt: Optional[datetime],
    end_dt: Optional[datetime],
) -> Tuple[Dict[str, Dict[str, Any]], List[str]]:
    reports: Dict[str, Dict[str, Any]] = {}
    missing: List[str] = []
    for printer in printers:
        report = load_joblog_report(printer, user_kw, mode_kw, computer_kw, start_dt, end_dt)
        if not report:
            missing.append(printer)
            continue
        report["printer"] = printer
        reports[printer] = report
    return reports, missing


def _format_usage_entries(
    stats: Dict[str, Dict[str, Any]],
    categories: List[str],
    limit: int,
    show_zero: bool,
) -> List[Dict[str, Any]]:
    ordered = []
    for record in stats.values():
        cat_values = [
            {
                "key": key,
                "label": USAGE_CATEGORY_CONFIG[key]["label"],
                "pages": record["totals"].get(key, 0),
            }
            for key in categories
        ]
        total = sum(item["pages"] for item in cat_values)
        if total == 0 and not show_zero:
            continue
        category_map = {item["key"]: item["pages"] for item in cat_values}
        ordered.append({
            "name": record["name"],
            "username": record.get("username", ""),  # Pass username through
            "total": total,
            "categories": cat_values,
            "category_map": category_map
        })

    ordered.sort(key=lambda item: item["total"], reverse=True)
    if limit > 0:
        ordered = ordered[:limit]
    return ordered


def _build_jobs_query() -> Dict[str, Any]:
    params = _parse_common_params(request.args)
    if params["per_page"] == 0:
        params["per_page"] = 3
    return params


def _prepare_jobs_context(query_args: Dict[str, Any]) -> Dict[str, Any]:
    # Parse pagination args (already parsed in _build_jobs_query)
    page = query_args.get("page", 1)
    per_page = query_args.get("per_page", 3)
    limit_val = query_args.get("limit", 5)
    
    # Simple bounds check
    if per_page < 1:
        per_page = 3
    query_args["per_page"] = per_page # Ensure context reflects effective value

    user_kw = query_args.get("user", "").strip()
    printer_pick = query_args.get("printer", "all")
    mode_pick = query_args.get("mode", "").strip()
    computer_pick = query_args.get("computer", "").strip()
    filename_pick = query_args.get("filename", "").strip()
    time_mode = query_args.get("time_mode", "all")

    # Time range
    start_dt, end_dt = None, None
    mode_display = ""

    if time_mode == "month":
        month_str = query_args.get("month", "")
        start_dt, end_dt = parse_month_range(month_str)
        mode_display = f"月份: {month_str}" if month_str else ""
    elif time_mode == "week":
        week_str = query_args.get("week", "")
        start_dt, end_dt = parse_week_range(week_str)
        mode_display = f"週次: {week_str}" if week_str else ""
    elif time_mode == "custom":
        start_str = query_args.get("start", "")
        end_str = query_args.get("end", "")
        start_dt = parse_time_value(start_str)
        end_dt = parse_time_value(end_str)
        mode_display = "自訂時間"

    # 1. Fetch paginated users (aggregated stats for sorting)
    users_list, total_users = fetch_aggregated_users_paginated(
        page, per_page, 
        printer_addr=printer_pick,
        user_kw=user_kw,
        mode_kw=mode_pick,
        computer_kw=computer_pick,
        start_dt=start_dt,
        end_dt=end_dt,
        filename_kw=filename_pick
    )

    if not users_list:
        errors = []
        if user_kw or mode_display or filename_pick:
            errors = ["查無符合條件的資料。"]
        return {
            "query": query_args,
            "errors": errors,
            "results": [],
            "query_string": urllib.parse.urlencode(query_args),
            "pagination": {
                "page": page,
                "per_page": per_page,
                "total_users": 0,
                "total_pages": 0
            }
        }

    # 2. Fetch detailed logs for these users
    detailed_entries = fetch_job_logs_by_users(
        users_list,
        printer_addr=printer_pick,
        mode_kw=mode_pick,
        computer_kw=computer_pick,
        start_dt=start_dt,
        end_dt=end_dt,
        filename_kw=filename_pick
    )
    
    # 3. Aggregate into report format
    user_map = {} # (user, login) -> list of entries
    for entry in detailed_entries:
        key = (entry["user"], entry["login"])
        if key not in user_map:
            user_map[key] = []
        user_map[key].append(entry)
        
    final_blocks = []
    # Preserve order from users_list (which is sorted by pages DESC)
    for u in users_list:
        key = (u["user"], u["login"])
        entries = user_map.get(key, [])
        if not entries:
            continue
            
        # Calculate totals
        total_jobs = len(entries)
        total_pages = sum(e["pages"] for e in entries)
        total_bw = sum(e["bw"] for e in entries)
        total_color = sum(e["color"] for e in entries)
        
        # Printer sub-totals
        p_stats = defaultdict(lambda: {"jobs": 0, "pages": 0})
        for e in entries:
            p_addr = e["printer"]
            p_stats[p_addr]["jobs"] += 1
            p_stats[p_addr]["pages"] += e["pages"]
            
        p_summaries = []
        for addr, st in p_stats.items():
            label = host_tag(addr)
            p_summaries.append({"label": label, "jobs": st["jobs"], "pages": st["pages"]})
        p_summaries.sort(key=lambda x: x["pages"], reverse=True)

        user_display = normalize_name(u["user"], "未知")
        login_display = normalize_name(u["login"], "N/A")

        # Slice entries for display if limit is set
        display_entries = entries
        if limit_val > 0:
            display_entries = entries[:limit_val]

        final_blocks.append({
            "name": user_display,
            "login": login_display,
            "totals": {
                "jobs": total_jobs,
                "pages": total_pages,
                "bw": total_bw,
                "color": total_color
            },
            "printer_totals": p_summaries,
            "entries": display_entries
        })

    # Calc total pages for pagination
    import math
    total_pages_count = math.ceil(total_users / per_page) if per_page > 0 else 0

    # Calculate global total jobs count for display
    total_jobs_count = fetch_total_jobs_count(
        printer_addr=printer_pick,
        user_kw=user_kw,
        mode_kw=mode_pick,
        computer_kw=computer_pick,
        start_dt=start_dt,
        end_dt=end_dt,
        filename_kw=filename_pick
    )

    return {
        "query": query_args,
        "errors": [],
        "results": final_blocks,
        "query_string": urllib.parse.urlencode(query_args),
        "pagination": {
            "page": page,
            "per_page": per_page,
            "total_users": total_users,
            "total_jobs": total_jobs_count,
            "total_pages": total_pages_count,
            "has_prev": page > 1,
            "has_next": page < total_pages_count,
            "prev_num": page - 1,
            "next_num": page + 1
        }
    }





def _build_counts_query() -> Dict[str, Any]:
    params = _parse_common_params(request.args)
    # Add counts-specific fields
    params.update({
        "categories": request.args.getlist("category"),
        "show_zero": request.args.get("show_zero") == "on",
        "view_mode": request.args.get("view_mode", "single_printer"),
        "export_scope": request.args.get("export_scope", "filtered"),
    })
    # Default per_page for counts is 20 if not specified (0 from parser)
    if params["per_page"] == 0:
        params["per_page"] = 20
    return params



def _prepare_counts_context(query: Dict[str, Any], export_mode: bool = False) -> Dict[str, Any]:
    # Parse pagination
    if export_mode:
        page = 1
        per_page = 0
    else:
        page = query.get("page", 1)
        per_page = query.get("per_page", 20)
        
        if per_page not in [10, 20, 30, 50, 100]:
            per_page = 20
    
    # Update query dict so template receives the effective value
    query["per_page"] = per_page # Keep as int or str? Template handles both usually, but strict equals might vary.
                                 # _parse_common_params returns int.
                                 # Let's keep it consistent.

    limit = query.get("limit", 0)


    errors: List[str] = []
    start_dt, end_dt = _resolve_time_range_from_query(
        query["time_mode"], query["month"], query["week"], query["start"], query["end"], errors
    )
    printer_pick = query["printer"]
    user_kw = query["user"] or None
    requested_categories = query["categories"] or DEFAULT_USAGE_CATEGORIES
    categories = [key for key in requested_categories if key in USAGE_CATEGORY_CONFIG] or DEFAULT_USAGE_CATEGORIES
    view_mode = query.get("view_mode", "single_printer")
    
    category_columns = [
        {"key": key, "label": USAGE_CATEGORY_CONFIG[key]["label"]}
        for key in categories
    ]
    
    results = []
    aggregated = None
    total_users = 0
    
    # View Mode Logic
    if view_mode == "single_printer":
        # Single Printer View: Show only one printer's data
        # If printer_pick is "all", default to first printer
        if printer_pick == "all" and PRINTERS:
            printer_pick = PRINTERS[0]
            query["printer"] = printer_pick # Sync back to query for UI
        
        if printer_pick != "all":
            # Fetch paginated users for selected printer
            p_users, total_users = fetch_aggregated_users_paginated(
                page, per_page,
                printer_addr=printer_pick,
                user_kw=user_kw,
                mode_kw=None,
                computer_kw=None,
                start_dt=start_dt,
                end_dt=end_dt
            )
            
            if p_users:
                p_detailed = fetch_job_logs_by_users(
                    p_users,
                    printer_addr=printer_pick,
                    mode_kw=None,
                    computer_kw=None,
                    start_dt=start_dt,
                    end_dt=end_dt
                )
                stats = aggregate_usage_by_categories(p_detailed, categories, None)
                formatted_entries = _format_usage_entries(stats, categories, 0, False)
                
                results.append({
                    "printer": printer_pick,
                    "label": _printer_label(printer_pick),
                    "file_name": "DB_Query",
                    "entries": formatted_entries,
                })
            else:
                results.append({
                    "printer": printer_pick,
                    "label": _printer_label(printer_pick),
                    "file_name": "DB_Query",
                    "entries": [],
                })
    
    elif view_mode == "all_printers":
        # All Printers View: Unified table with printer column
        # Fetch paginated users across ALL printers
        all_users, total_users = fetch_aggregated_users_paginated(
            page, per_page,
            printer_addr="all",  # All printers
            user_kw=user_kw,
            mode_kw=None,
            computer_kw=None,
            start_dt=start_dt,
            end_dt=end_dt
        )
        
        if all_users:
            # Calculate total records (user-printer pairs) for display
            total_records = fetch_total_user_printer_pairs(
                printer_addr="all",
                user_kw=user_kw,
                mode_kw=None,
                computer_kw=None,
                start_dt=start_dt,
                end_dt=end_dt
            )
            
            # Fetch detailed logs for these users across all printers
            all_detailed = fetch_job_logs_by_users(
                all_users,
                printer_addr="all",
                mode_kw=None,
                computer_kw=None,
                start_dt=start_dt,
                end_dt=end_dt
            )
            
            # Group logs by printer first, then aggregate per printer
            from collections import defaultdict
            printer_logs = defaultdict(list)
            for log in all_detailed:
                printer = log.get("printer", "")  # Changed from printer_addr to printer
                printer_logs[printer].append(log)
            
            # Aggregate each printer's data and combine with printer info
            unified_entries = []
            for printer, logs in printer_logs.items():
                # Use existing aggregation function for consistency
                stats = aggregate_usage_by_categories(logs, categories, None)
                
                # stats is a dict: {user_key: {"name": ..., "totals": {...}, "total": ...}}
                for user_key, user_data in stats.items():
                    category_map = {}
                    for cat in categories:
                        # user_data["totals"] contains the category counts
                        category_map[cat] = user_data["totals"].get(cat, 0)
                    
                    unified_entries.append({
                        "name": user_data["name"],
                        "username": user_data.get("username", ""),  # Add username
                        "login": "",  # Not available in aggregated stats
                        "category_map": category_map,
                        "total": user_data["total"],
                        "printer": printer
                    })
            
            # Sort by total pages desc
            unified_entries.sort(key=lambda x: x["total"], reverse=True)
            
            # Store in results (single "block" for all printers)
            results = unified_entries
    
    elif view_mode == "aggregated":
        # Aggregated View: Cross-printer user totals
        agg_users, total_users = fetch_aggregated_users_paginated(
            page, per_page,
            printer_addr="all",
            user_kw=user_kw,
            mode_kw=None,
            computer_kw=None,
            start_dt=start_dt,
            end_dt=end_dt
        )
        
        if agg_users:
            agg_detailed = fetch_job_logs_by_users(
                agg_users,
                printer_addr="all",
                mode_kw=None,
                computer_kw=None,
                start_dt=start_dt,
                end_dt=end_dt
            )
            agg_stats = aggregate_usage_by_categories(agg_detailed, categories, None)
            aggregated_entries = _format_usage_entries(agg_stats, categories, 0, False)
            aggregated = {"entries": aggregated_entries} if aggregated_entries else None
    
    # Calculate total pages
    import math
    total_pages_count = math.ceil(total_users / per_page) if per_page > 0 else 0

    return {
        "results": results,
        "errors": errors,
        "category_columns": category_columns,
        "reports": [],  # unused
        "aggregated": aggregated,
        "categories": categories,
        "view_mode": view_mode,
        "pagination": {
            "page": page,
            "per_page": per_page,
            "total_users": total_users,
            "total_records": total_records if view_mode == "all_printers" and locals().get("total_records") else 0,
            "total_pages": total_pages_count,
            "has_prev": page > 1,
            "has_next": page < total_pages_count,
            "prev_num": page - 1,
            "next_num": page + 1
        }
    }


def _build_leaders_query() -> Dict[str, Any]:
    return {
        "printer": request.args.get("printer", "all"),
        "user": request.args.get("user", "").strip(),
        "mode": request.args.get("mode", "").strip(),
        "computer": request.args.get("computer", "").strip(),
        "time_mode": request.args.get("time_mode", "all"),
        "month": request.args.get("month", "").strip(),
        "week": request.args.get("week", "").strip(),
        "start": request.args.get("start", "").strip(),
        "end": request.args.get("end", "").strip(),
        "page": request.args.get("page", "1"),
        "per_page": request.args.get("per_page", "20"),
        "view_mode": request.args.get("view_mode", "all_printers"),
    }


def _prepare_leaders_context(query: Dict[str, Any]) -> Dict[str, Any]:
    view_mode = query.get("view_mode", "all_printers")
    
    # Pagination for aggregated list
    page = _to_int(query["page"], 1)
    if page < 1: page = 1
    per_page = _to_int(query["per_page"], 20)
    if per_page < 5: per_page = 5
    
    errors: List[str] = []
    start_dt, end_dt = _resolve_time_range_from_query(
        query["time_mode"], query["month"], query["week"], query["start"], query["end"], errors
    )
    
    user_kw = query["user"] or None
    mode_kw = query["mode"] or None
    computer_kw = query["computer"] or None
    
    # Determine which printers to fetch based on view mode
    if view_mode == "single_printer":
        # Single printer view: use selected printer (or first if "all")
        selected_printer = query["printer"]
        if selected_printer == "all" or not selected_printer:
            printers = [PRINTERS[0]] if PRINTERS else []
        else:
            printers = [selected_printer]
    else:
        # all_printers or aggregated: fetch all printers
        printers = _selected_printers("all")

    reports, _ = _collect_job_reports(printers, user_kw, mode_kw, computer_kw, start_dt, end_dt)

    # Collect all rows based on view mode
    all_rows = []
    show_printer_column = False
    
    if view_mode == "aggregated":
        # Aggregated Mode: Group by User
        if reports:
            summary = aggregate_joblog_reports(list(reports.values()))
            if summary:
                # summary["users"] is already sorted by pages desc in aggregate_joblog_reports (but we ensure it)
                all_rows = sorted(summary["users"], key=lambda u: u["pages"], reverse=True)
                # Aggregated rows structure: {user, login, jobs, bw, color, pages}
    else:
        # Detail Mode: All Printers or Single Printer
        # Flatten all top_users from all reports into a single list
        for printer in printers:
            report = reports.get(printer)
            if not report:
                continue
                
            label = _printer_label(printer)
            for user_entry in report["top_users"]:
                # Create a copy to enrich with printer data
                row = user_entry.copy()
                row["printer"] = printer
                row["printer_label"] = label
                all_rows.append(row)
        
        # Sort globally by pages descending
        all_rows.sort(key=lambda u: u["pages"], reverse=True)
        
        # Show printer column only in "all_printers" mode
        if view_mode == "all_printers":
            show_printer_column = True

    # Global Pagination
    total_users = len(all_rows)
    # Calculate unique users count for "all_printers" view where one user may appear multiple times
    total_unique_users = len(set((r.get('user') or '').lower() + (r.get('login') or '').lower() for r in all_rows))
    
    import math
    total_pages_count = math.ceil(total_users / per_page) if per_page > 0 else 0
    
    start_idx = (page - 1) * per_page
    end_idx = start_idx + per_page
    display_rows = all_rows[start_idx:end_idx]
    
    pagination = {
        "page": page,
        "per_page": per_page,
        "total_users": total_users,
        "total_unique_users": total_unique_users,
        "total_pages": total_pages_count,
        "has_prev": page > 1,
        "prev_num": page - 1,
        "has_next": page < total_pages_count,
        "next_num": page + 1
    }

    # Calculate Totals for the current selection (all fetched reports)
    # Note: For aggregated, summary["totals"] exists. For detail, we sum up reports.
    grand_totals = {"jobs": 0, "bw": 0, "color": 0, "pages": 0}
    if view_mode == "aggregated" and reports:
        # Use existing summary totals if available
        # Need to re-fetch summary or calculating it? We calculated it above.
        # Ideally we refactor code to keep summary variable accessible.
        pass # We'll calculate manually below to be safe and consistent
        
    for r in reports.values():
        t = r["totals"]
        grand_totals["jobs"] += t["jobs"]
        grand_totals["bw"] += t["bw"]
        grand_totals["color"] += t["color"]
        grand_totals["pages"] += t["pages"]

    return {
        "rows": display_rows,
        "show_printer_column": show_printer_column,
        "totals": grand_totals,
        "errors": errors,
        "pagination": pagination,
        "view_mode": view_mode
    }


def _workbook_response(wb: Workbook, filename: str):
    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    return send_file(
        stream,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


def _build_jobs_workbook(reports: List[Dict[str, Any]]) -> Workbook:
    wb = Workbook()
    first_sheet = True
    for report in reports:
        printer = report["printer"]
        title = _printer_label(printer)[:31] or "Sheet"
        ws = wb.active if first_sheet else wb.create_sheet(title)
        ws.title = title
        first_sheet = False
        ws.append(["工作ID", "開始時間", "模式", "用戶", "登入名稱", "電腦名稱", "黑白張數", "彩色張數", "總張數"])
        for entry in report.get("entries", []):
            ws.append(
                [
                    entry.get("job_id") or entry.get("account_job_id") or "?",
                    format_dt(entry.get("start")),
                    entry.get("mode") or "N/A",
                    entry.get("user_display") or entry.get("user") or "未知",
                    entry.get("login_display") or entry.get("login") or "N/A",
                    entry.get("computer") or "N/A",
                    entry.get("bw", 0),
                    entry.get("color", 0),
                    entry.get("pages", 0),
                ]
            )
    return wb


def _build_counts_workbook(results: List[Dict[str, Any]], categories: List[str]) -> Workbook:
    wb = Workbook()
    first_sheet = True
    for block in results:
        title = _printer_label(block["printer"])[:31]
        ws = wb.active if first_sheet else wb.create_sheet(title)
        ws.title = title
        first_sheet = False
        headers = ["用戶", "帳號"] + [USAGE_CATEGORY_CONFIG[key]["label"] for key in categories] + ["總張數"]
        ws.append(headers)
        for item in block.get("entries", []):
            category_map = item.get("category_map", {})
            username = item.get("username", "")
            display_name = ldap_service.get_user_display_name(username) if username else item["name"]
            row = [display_name, username]
            for key in categories:
                row.append(category_map.get(key, 0))
            row.append(item["total"])
            ws.append(row)
    return wb








def _build_combined_counts_workbook(entries: List[Dict[str, Any]], categories: List[str]) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "跨機器彙總"
    headers = ["用戶", "帳號"] + [USAGE_CATEGORY_CONFIG[key]["label"] for key in categories] + ["總張數"]
    ws.append(headers)
    for item in entries:
        category_map = item.get("category_map", {})
        row = [item.get("name", "未知"), item.get("username", "")]
        for key in categories:
            row.append(category_map.get(key, 0))
        row.append(item.get("total", 0))
        ws.append(row)
    return wb


def _build_all_printers_workbook(entries: List[Dict[str, Any]], categories: List[str]) -> Workbook:
    """
    Build workbook for 'all_printers' view mode.
    Creates a single sheet with unified table including printer column.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "所有列印機統計"
    
    # Headers: User, Username, Categories..., Total, Printer
    headers = ["用戶", "帳號"] + [USAGE_CATEGORY_CONFIG[key]["label"] for key in categories] + ["總張數", "列印機"]
    ws.append(headers)
    
    # Add data rows
    for item in entries:
        category_map = item.get("category_map", {})
        username = item.get("username", "")
        display_name = ldap_service.get_user_display_name(username) if username else item.get("name", "未知")
        row = [display_name, username]
        for key in categories:
            row.append(category_map.get(key, 0))
        row.append(item.get("total", 0))
        row.append(_printer_label(item.get("printer", "")))
        ws.append(row)
    
    return wb



def _build_leaders_workbook(rows: List[Dict[str, Any]], show_printer_column: bool) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "排行榜"
    
    headers = ["用戶", "登入名稱", "筆數", "黑白張數", "彩色張數", "總張數"]
    if show_printer_column:
        headers.append("列印機")
        
    ws.append(headers)
    
    for item in rows:
        row_data = [
            item["user"],
            item["login"],
            item["jobs"],
            item["bw"],
            item["color"],
            item["pages"],
        ]
        if show_printer_column:
            row_data.append(item.get("printer_label", ""))
            
        ws.append(row_data)

    return wb


@app.route("/")
def index():
    # Fetch recent update logs (top 5) for dashboard
    # Fetch recent update logs (paginated)
    import sharp_mfp_export
    import math
    
    page = request.args.get('log_page', 1, type=int)
    per_page = 5
    logs = []
    total_logs = 0
    total_pages = 0
    
    try:
        conn = sharp_mfp_export.get_db_connection()
        with conn.cursor() as cursor:
            # Get total count
            cursor.execute("SELECT COUNT(*) as c FROM update_logs")
            total_logs = cursor.fetchone()['c']
            
            # Calculate offset
            total_pages = math.ceil(total_logs / per_page)
            if page < 1: page = 1
            if page > total_pages and total_pages > 0: page = total_pages
            
            offset = (page - 1) * per_page
            
            # Fetch paginated logs (Construct SQL manually to avoid LIMIT '5' syntax error)
            sql = f"SELECT * FROM update_logs ORDER BY id DESC LIMIT {int(per_page)} OFFSET {int(offset)}"
            cursor.execute(sql)
            logs = cursor.fetchall()
        conn.close()
    except Exception as e:
        import logging
        logging.error(f"Failed to fetch logs: {e}")
        pass # Fail gracefully if DB not ready
        
    return render_template("index.html", 
                         printer_choices=PRINTER_CHOICES, 
                         logs=logs,
                         current_page=page,
                         total_pages=total_pages)


@app.route("/counts")
@cache.cached(timeout=300, query_string=True)
def counts():
    query = _build_counts_query()
    context = _prepare_counts_context(query)
    query_string = request.query_string.decode() if request.query_string else ""
    return render_template(
        "counts.html",
        printer_choices=PRINTER_CHOICES,
        category_choices=CATEGORY_CHOICES,
        query=query,
        selected_categories=context["categories"],
        results=context["results"],
        errors=context["errors"],
        category_columns=context["category_columns"],
        aggregated=context["aggregated"],
        query_string=query_string,
        pagination=context.get("pagination"),
    )


@app.route("/jobs")
def jobs():
    query = _build_jobs_query()
    context = _prepare_jobs_context(query)
    query_string = request.query_string.decode() if request.query_string else ""
    return render_template(
        "jobs.html",
        printer_choices=PRINTER_CHOICES,
        query=context["query"],
        results=context["results"],
        errors=context.get("errors", []),
        query_string=query_string,
        pagination=context.get("pagination"),
    )


@app.route("/leaders")
@cache.cached(timeout=300, query_string=True)
def leaders():
    query = _build_leaders_query()
    context = _prepare_leaders_context(query)
    query_string = request.query_string.decode() if request.query_string else ""
    return render_template(
        "leaders.html",
        printer_choices=PRINTER_CHOICES,
        query=query,
        rows=context["rows"],
        show_printer_column=context["show_printer_column"],
        totals=context["totals"],
        errors=context["errors"],
        query_string=query_string,
        pagination=context.get("pagination"),
        view_mode=context.get("view_mode", "all_printers"),
    )


def _build_user_jobs_workbook(user_blocks: List[Dict[str, Any]]) -> Workbook:
    """
    Build workbook for jobs page export (user-grouped data).
    Each user_block has: name, login, totals, entries
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "作業紀錄"
    
    # Header row
    ws.append(["用戶名稱", "登入名稱", "工作ID", "開始時間", "模式", "電腦名稱", "列印機", "黑白張數", "彩色張數", "總張數", "檔案名稱"])
    
    for block in user_blocks:
        user_name = block.get("name", "未知")
        login_name = block.get("login", "N/A")
        
        for entry in block.get("entries", []):
            ws.append([
                user_name,
                login_name,
                entry.get("job_id") or entry.get("account_job_id") or "?",
                format_dt(entry.get("start")),
                entry.get("mode") or "N/A",
                entry.get("computer") or "N/A",
                _printer_label(entry.get("printer", "")),
                entry.get("bw", 0),
                entry.get("color", 0),
                entry.get("pages", 0),
                entry.get("file_name") or ""
            ])
    
    return wb


@app.route("/export/jobs")
def export_jobs():
    query = _build_jobs_query()
    context = _prepare_jobs_context(query)
    
    # Get user blocks from context
    user_blocks = context.get("results", [])
    
    wb = _build_user_jobs_workbook(user_blocks)
    return _workbook_response(wb, "jobs_export.xlsx")


@app.route("/export/stats")
def export_stats():
    """
    Unified export route that handles all three view modes and export scopes.
    - view_mode: single_printer, all_printers, aggregated
    - export_scope: filtered (current page/filter), all (all matching data)
    """
    query = _build_counts_query()
    view_mode = query.get("view_mode", "single_printer")
    export_scope = query.get("export_scope", "filtered")
    categories = query.get("categories", [])
    
    # Prepare context based on export scope
    if export_scope == "all":
        # Export ALL data: create clean query without any filters
        clean_query = {
            "view_mode": view_mode,
            "categories": categories,
            "printer": "all",  # All printers
            "user": "",  # No user filter
            "time_mode": "all",  # All time
            "month": "",
            "week": "",
            "start": "",  # Changed from None to empty string
            "end": "",    # Changed from None to empty string
            "page": 1,
            "per_page": 0,  # No pagination
            "limit": "0",   # Missing key caused 500
            "export_scope": "all"
        }
        context = _prepare_counts_context(clean_query, export_mode=True)
    else:
        # Export filtered data: use current filters + current page
        context = _prepare_counts_context(query, export_mode=False)
    
    categories = context["categories"]
    
    # Build workbook based on view mode
    if view_mode == "single_printer":
        # Single printer: export per-printer workbook
        wb = _build_counts_workbook(context["results"], categories)
        filename = "stats_single_printer.xlsx"
    
    elif view_mode == "all_printers":
        # All printers: export unified table with printer column
        # Need to create a new workbook builder for this format
        wb = _build_all_printers_workbook(context["results"], categories)
        filename = "stats_all_printers.xlsx"
    
    elif view_mode == "aggregated":
        # Aggregated: export cross-printer summary
        aggregated = context.get("aggregated") or {"entries": []}
        entries = aggregated.get("entries", [])
        wb = _build_combined_counts_workbook(entries, categories)
        filename = "stats_aggregated.xlsx"
    
    else:
        # Fallback to single printer mode
        wb = _build_counts_workbook(context["results"], categories)
        filename = "stats_export.xlsx"
    
    return _workbook_response(wb, filename)


# Keep legacy route for backward compatibility (redirects to new unified route)
@app.route("/export/stats_combined")
def export_stats_combined():
    """Legacy route: redirects to unified export with aggregated view mode"""
    from flask import redirect, url_for
    # Preserve query parameters and add view_mode=aggregated
    args = request.args.to_dict(flat=False)
    args['view_mode'] = ['aggregated']
    return redirect(url_for('export_stats', **{k: v[0] if len(v) == 1 else v for k, v in args.items()}))


@app.route("/export/leaders")
def export_leaders():
    export_range = request.args.get("export_range", "current_filter")
    
    query = _build_leaders_query()
    # Use large per_page to ensure all rows are returned (no pagination for export)
    query["per_page"] = "1000000"
    query["page"] = "1"
    
    # If exporting all data, clear all filters
    if export_range == "all_data":
        query["user"] = ""
        query["mode"] = ""
        query["computer"] = ""
        query["time_mode"] = "all"
        query["month"] = ""
        query["week"] = ""
        query["start"] = ""
        query["end"] = ""
    
    context = _prepare_leaders_context(query)
    wb = _build_leaders_workbook(context["rows"], context["show_printer_column"])
    
    # Add descriptive filename
    if export_range == "all_data":
        filename = "leaders_all_data.xlsx"
    else:
        filename = "leaders_filtered.xlsx"
    
    return _workbook_response(wb, filename)




if __name__ == "__main__":
    try:
        from waitress import serve
        print("Starting Waitress production server on http://0.0.0.0:5000")
        serve(app, host="0.0.0.0", port=5000, threads=8)
    except ImportError:
        print("Waitress not installed, falling back to Flask development server")
        app.run(host="0.0.0.0", port=5000, debug=False)

